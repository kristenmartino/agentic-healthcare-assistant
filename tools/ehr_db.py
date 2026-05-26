"""Electronic Health Records (EHR) database wrapper.

Loads `records.xlsx` (instructor-provided) into a cleaned SQLite table.
Exposes simple CRUD helpers used by the records and history nodes.

Cleanup pass on the source xlsx:
- Deduplicate rows that share normalized phone (e.g., Rebeca Nagle appears 3x).
- Normalize Phone_number into a digits-only form for matching.
- Generate a stable patient_id = SHA1(phone || age) prefix-12.
- Replace empty/blank Email/Summary/Address with NULL.

This script is idempotent: re-running it overwrites the patients table.
"""
from __future__ import annotations

import hashlib
import logging
import re
import sqlite3
from collections.abc import Iterator
from contextlib import contextmanager
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)


def _normalize_phone(raw: Any) -> str:
    """Strip everything except digits."""
    if raw is None:
        return ""
    return re.sub(r"[^\d]", "", str(raw))


def _patient_id(phone_digits: str, age: Any, name: str) -> str:
    """Deterministic 12-char ID. Falls back to name+age if no phone."""
    seed = phone_digits or f"{name.lower().strip()}|{age}"
    return hashlib.sha1(seed.encode()).hexdigest()[:12]


def _coerce_age(raw: Any) -> int | None:
    if raw is None:
        return None
    try:
        return int(raw)
    except (TypeError, ValueError):
        return None


def _is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


@contextmanager
def _connect(db_path: str) -> Iterator[sqlite3.Connection]:
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    try:
        yield conn
        conn.commit()
    finally:
        conn.close()


SCHEMA = """
CREATE TABLE IF NOT EXISTS patients (
    patient_id TEXT PRIMARY KEY,
    name TEXT NOT NULL,
    phone_normalized TEXT,
    phone_raw TEXT,
    email TEXT,
    age INTEGER,
    gender TEXT,
    address TEXT,
    summary TEXT,
    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
);

CREATE INDEX IF NOT EXISTS idx_patients_name ON patients(name COLLATE NOCASE);
CREATE INDEX IF NOT EXISTS idx_patients_phone ON patients(phone_normalized);
"""


def initialize_ehr(records_xlsx_path: str, db_path: str) -> dict[str, int]:
    """Load records.xlsx into a cleaned SQLite table. Returns counts."""
    Path(db_path).parent.mkdir(parents=True, exist_ok=True)

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is required. Install with: pip install openpyxl"
        ) from exc

    if not Path(records_xlsx_path).exists():
        raise FileNotFoundError(f"records.xlsx not found at {records_xlsx_path}")

    wb = load_workbook(records_xlsx_path, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise RuntimeError("records.xlsx is empty")

    # First row is headers — be tolerant of column order
    headers = [str(h or "").strip() for h in rows[0]]

    def col(record: dict, *names: str) -> Any:
        for n in names:
            if n in record and not _is_blank(record[n]):
                return record[n]
        return None

    seen_ids: set[str] = set()
    cleaned: list[dict] = []
    raw_count = 0
    duplicate_count = 0
    skipped_count = 0

    for r in rows[1:]:
        raw_count += 1
        record = dict(zip(headers, r, strict=False))

        name = col(record, "Name")
        if not name:
            skipped_count += 1
            continue
        name = str(name).strip()

        phone_raw = col(record, "Phone_number", "Phone")
        phone_norm = _normalize_phone(phone_raw)
        age = _coerce_age(col(record, "Age"))

        pid = _patient_id(phone_norm, age, name)
        if pid in seen_ids:
            duplicate_count += 1
            continue
        seen_ids.add(pid)

        cleaned.append({
            "patient_id": pid,
            "name": name,
            "phone_normalized": phone_norm or None,
            "phone_raw": str(phone_raw) if phone_raw else None,
            "email": str(col(record, "Email") or "") or None,
            "age": age,
            "gender": str(col(record, "Gender") or "") or None,
            "address": str(col(record, "Address") or "") or None,
            "summary": str(col(record, "Summary") or "") or None,
        })

    with _connect(db_path) as conn:
        conn.executescript(SCHEMA)
        # Wipe + reload (idempotent)
        conn.execute("DELETE FROM patients")
        conn.executemany(
            """
            INSERT INTO patients (patient_id, name, phone_normalized, phone_raw,
                                  email, age, gender, address, summary)
            VALUES (:patient_id, :name, :phone_normalized, :phone_raw,
                    :email, :age, :gender, :address, :summary)
            """,
            cleaned,
        )

    logger.info(
        "EHR seeded: %d rows in xlsx → %d unique patients (%d duplicates, %d skipped)",
        raw_count, len(cleaned), duplicate_count, skipped_count,
    )
    return {
        "raw_rows": raw_count,
        "unique_patients": len(cleaned),
        "duplicates_dropped": duplicate_count,
        "skipped_rows": skipped_count,
    }


def list_patients(db_path: str) -> list[dict]:
    """Return all patients. Used by the Streamlit sidebar selector."""
    with _connect(db_path) as conn:
        rows = conn.execute(
            "SELECT patient_id, name, age, gender, summary FROM patients ORDER BY name"
        ).fetchall()
    return [dict(r) for r in rows]


def find_patient_by_name(db_path: str, name: str) -> dict | None:
    """Case-insensitive name match. Returns the first hit or None."""
    if not name:
        return None
    with _connect(db_path) as conn:
        row = conn.execute(
            "SELECT * FROM patients WHERE name LIKE ? COLLATE NOCASE LIMIT 1",
            (f"%{name.strip()}%",),
        ).fetchone()
    return dict(row) if row else None


def add_or_update_patient(db_path: str, fields: dict) -> dict:
    """Insert or update a patient. Returns {operation, patient_id, before, after}.

    `fields` must include `name`. Other fields are optional.
    Operation: 'insert' if no matching patient_id existed, else 'update'.
    """
    name = (fields.get("name") or "").strip()
    if not name:
        raise ValueError("Cannot add/update a record without a name")

    phone_raw = fields.get("phone_raw") or fields.get("phone_number")
    phone_norm = _normalize_phone(phone_raw)
    age = _coerce_age(fields.get("age"))
    pid = fields.get("patient_id") or _patient_id(phone_norm, age, name)

    with _connect(db_path) as conn:
        before = conn.execute(
            "SELECT * FROM patients WHERE patient_id = ?", (pid,),
        ).fetchone()
        before_dict = dict(before) if before else None

        if before is None:
            conn.execute(
                """
                INSERT INTO patients (patient_id, name, phone_normalized, phone_raw,
                                      email, age, gender, address, summary)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    pid, name, phone_norm or None, phone_raw,
                    fields.get("email"), age, fields.get("gender"),
                    fields.get("address"), fields.get("summary"),
                ),
            )
            operation = "insert"
        else:
            # Merge: update only fields that were provided
            updates: list[str] = []
            values: list[Any] = []
            for col, key in [
                ("name", "name"), ("phone_raw", "phone_raw"),
                ("phone_normalized", "phone_normalized"),
                ("email", "email"), ("age", "age"),
                ("gender", "gender"), ("address", "address"),
                ("summary", "summary"),
            ]:
                if key in fields and fields[key] is not None:
                    updates.append(f"{col} = ?")
                    values.append(fields[key] if key != "phone_normalized" else phone_norm or None)
            if updates:
                updates.append("updated_at = CURRENT_TIMESTAMP")
                values.append(pid)
                conn.execute(
                    f"UPDATE patients SET {', '.join(updates)} WHERE patient_id = ?",
                    values,
                )
            operation = "update"

        after = conn.execute(
            "SELECT * FROM patients WHERE patient_id = ?", (pid,),
        ).fetchone()
        after_dict = dict(after) if after else None

    return {
        "operation": operation,
        "patient_id": pid,
        "before": before_dict,
        "after": after_dict,
    }


# CLI: python -m tools.ehr_db init <records.xlsx> <db_path>
if __name__ == "__main__":
    import sys
    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
    if len(sys.argv) >= 4 and sys.argv[1] == "init":
        result = initialize_ehr(sys.argv[2], sys.argv[3])
        print(result)
    else:
        print("Usage: python -m tools.ehr_db init <records.xlsx> <db_path>")
        sys.exit(1)
